#########################################################################

# THIS PROGRAM CLEANS UP UNNECCESSARY CLOUMNS AND ROWS FOR THE 
# PLAYER STATS FILES. WILL BE NECCESSARY FOR ML 

#########################################################################
from openpyxl import load_workbook
#########################################################################
# ASKS USER FOR LAST NAME OF PLAYER
while True:
    if input('Would you like to clean the player file? Enter y/n : ') != 'y':
        break
    else:
        # PLEASE CHOOSE WHICH FILE TYPE YOU'D LIKE TO USE 
        print('Which Player stats would you like to clean? ')
        last_name = input('Enter by Last Name: ')
        wb = load_workbook(fr"( ~ YOUR PATH HERE ~ )\{last_name}.xlsx")
        ws = wb.active
# DELETING ROWS AND COLUMNS MUST HAVE TO DELETE BOTTOM TO TOP
# OF EXCEL, IF NOT THEN IT WON'T COUNT THE DELETED ROWS FROM BEFORE
# AND WONT NEED AFTER RAN ONCE
        ws.delete_rows(1,2)
        ws.delete_rows(ws.max_row)
        ws.delete_cols(1)
        print("Cleaning up excel file...")
# SAVES THE NEW FILE 
# FILE MUST! BE CLOSED TO BE ABLE TO SAVE FILE
        wb.save(fr"( ~ YOUR PATH HERE ~ )\{last_name}.xlsx")
        print(' THE FILE HAS BEEN SAVED AND OVERWRITTEN!')
        break