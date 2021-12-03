#########################################################################

# PROGRAM WILL GRAB SPECIFIC PLAYER STATS AND WILL SAVE IN AN EXCEL FILE ->
# UNDER LASTNAME. EXCEL FILE WILL BE CLEANED (REMOVING UNNECCESSARY ROWS AND COLUMNS)
 
#########################################################################
from bs4 import BeautifulSoup
import pandas as pd
import requests
from openpyxl import load_workbook
#########################################################################
# ASKS USER FOR LAST NAME OF PLAYER
print('Which Player would you like to get stats for? ')
last_name = input('Enter by Last Name: ')

# GRABS THE PLAYER FROM THE WEBSITE
player_url = f"https://www.fftoday.com/stats/players?Search={last_name}"
page = requests.get(player_url).text
soup = BeautifulSoup(page, "html.parser")
############ PANDAS #####################################################
#GRAB TABLE FOR PLAYER CAREER STATS
print('Current stats for '+last_name+' are:')
dfs = pd.read_html(player_url)
df = dfs[7]
print(df)

# PRINTS THE PLAYER DATA INTO A EXCEL OR CSV FILE
df.to_excel(fr"C:\Users\Giovannie\Desktop\dataproject\playerstats\{last_name}.xlsx")
# # df.to_csv(r"C:\Users\Giovannie\Desktop\dataproject\tablescrape.csv")
print('PLayer Stats have been saved as '+last_name+'.xlsx file')
#########################################################################
# EXCEL CLEAN UP 
wb = load_workbook(fr"C:\Users\Giovannie\Desktop\dataproject\playerstats\{last_name}.xlsx")
ws = wb.active
# DELETING ROWS AND COLUMNS 
# MUST HAVE TO DELETE BOTTOM TO TOP OF EXCEL
# IF NOT THEN IT WON'T COUNT THE DELETED ROWS FROM BEFORE AND WONT NEED AFTER RAN ONCE
ws.delete_rows(1,2)
ws.delete_rows(ws.max_row)
ws.delete_cols(1)
print("Cleaning up excel file...")
# SAVES THE NEW FILE 
# FILE MUST! BE CLOSED TO BE ABLE TO SAVE FILE
wb.save(fr"C:\Users\Giovannie\Desktop\dataproject\playerstats\{last_name}.xlsx")
print(' THE FILE HAS BEEN SAVED AND OVERWRITTEN!')

#########################################################################